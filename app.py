import os
import io
import re
import csv
import sqlite3
from contextlib import closing
from datetime import date, datetime, timedelta, time
from zoneinfo import ZoneInfo

from telegram import Update, InputFile
from telegram.ext import ApplicationBuilder, CommandHandler, MessageHandler, ContextTypes, filters

# ---------- Config ----------
BOT_TOKEN = os.getenv("BOT_TOKEN", "")
ADMIN_IDS = {int(x) for x in os.getenv("ADMIN_IDS", "").replace(" ", "").split(",") if x.strip().isdigit()}
DB_PATH = os.getenv("DB_PATH", "ippt.db")  # set DB_PATH=/data/ippt.db on Railway
TZ_NAME = os.getenv("TZ", "Asia/Singapore")
try:
    TZINFO = ZoneInfo(TZ_NAME)
except Exception:
    TZINFO = ZoneInfo("UTC")

WINDOW_DAYS = 100
REMINDER_INTERVAL_DAYS = int(os.getenv("REMINDER_INTERVAL_DAYS", "10"))

# ---------- DB ----------
def db_connect():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    with conn:
        conn.execute("PRAGMA foreign_keys = ON")
    return conn

def _ensure_column(conn, table, column):
    cur = conn.cursor()
    cur.execute(f"PRAGMA table_info({table})")
    cols = [r[1] for r in cur.fetchall()]
    if column not in cols:
        cur.execute(f"ALTER TABLE {table} ADD COLUMN {column} TEXT")
        conn.commit()

def init_db():
    with closing(db_connect()) as conn:
        cur = conn.cursor()
        cur.execute("""
            CREATE TABLE IF NOT EXISTS personnel (
              personnel_id TEXT PRIMARY KEY,
              birthday     TEXT NOT NULL,    -- YYYY-MM-DD
              group_name   TEXT
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS users (
              telegram_id     INTEGER PRIMARY KEY,
              personnel_id    TEXT UNIQUE,
              verified_at     TEXT,
              completed_year  INTEGER,
              completed_at    TEXT,
              full_name       TEXT,
              FOREIGN KEY (personnel_id) REFERENCES personnel(personnel_id)
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS completions (
              id            INTEGER PRIMARY KEY AUTOINCREMENT,
              telegram_id   INTEGER NOT NULL,
              year          INTEGER NOT NULL,
              completed_at  TEXT NOT NULL
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS deferments (
              id           INTEGER PRIMARY KEY AUTOINCREMENT,
              personnel_id TEXT NOT NULL,
              year         INTEGER NOT NULL,
              reason       TEXT,
              status       TEXT CHECK (status IN ('approved')) DEFAULT 'approved',
              created_at   TEXT NOT NULL,
              UNIQUE (personnel_id, year),
              FOREIGN KEY (personnel_id) REFERENCES personnel(personnel_id)
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS cycle_notes (
              id           INTEGER PRIMARY KEY AUTOINCREMENT,
              personnel_id TEXT NOT NULL,
              year         INTEGER NOT NULL,    -- cycle start year
              reason       TEXT,
              created_at   TEXT NOT NULL,
              UNIQUE (personnel_id, year),
              FOREIGN KEY (personnel_id) REFERENCES personnel(personnel_id)
            )
        """)
        _ensure_column(conn, "users", "full_name")
        conn.commit()

# ---------- Date helpers ----------
def format_date(d: date) -> str:
    return d.strftime("%Y-%m-%d")

def parse_date_strict(s: str) -> date:
    return datetime.strptime(s, "%Y-%m-%d").date()

def adjusted_birthday_for_year(bday: date, year: int) -> date:
    try:
        return date(year, bday.month, bday.day)
    except ValueError:
        if bday.month == 2 and bday.day == 29:
            return date(year, 2, 28)
        raise

def current_local_date() -> date:
    return datetime.now(TZINFO).date()

def cycle_for_date(bday: date, on: date):
    start = adjusted_birthday_for_year(bday, on.year)
    if on < start:
        start = adjusted_birthday_for_year(bday, on.year - 1)
    end_excl = adjusted_birthday_for_year(bday, start.year + 1)
    return start, end_excl

def today_in_window(bday: date, today: date):
    start = adjusted_birthday_for_year(bday, today.year)
    end = start + timedelta(days=WINDOW_DAYS)
    if start <= today <= end:
        return True, start, end
    prev_start = adjusted_birthday_for_year(bday, today.year - 1)
    prev_end = prev_start + timedelta(days=WINDOW_DAYS)
    if prev_start <= today <= prev_end:
        return True, prev_start, prev_end
    return False, start, end

def window_for_date(bday: date, on: date):
    start = adjusted_birthday_for_year(bday, on.year)
    end = start + timedelta(days=WINDOW_DAYS)
    if start <= on <= end:
        return start, end
    prev_start = adjusted_birthday_for_year(bday, on.year - 1)
    prev_end = prev_start + timedelta(days=WINDOW_DAYS)
    if prev_start <= on <= prev_end:
        return prev_start, prev_end
    next_start = adjusted_birthday_for_year(bday, on.year + 1)
    next_end = next_start + timedelta(days=WINDOW_DAYS)
    return next_start, next_end

def iso_from_local_date(d: date, hour: int = 9, minute: int = 0) -> str:
    dt = datetime.combine(d, time(hour=hour, minute=minute, tzinfo=TZINFO))
    return dt.isoformat()

def is_admin(tid: int) -> bool:
    return tid in ADMIN_IDS

# ---------- DB helpers ----------
def get_personnel_and_user(conn: sqlite3.Connection, telegram_id: int):
    cur = conn.cursor()
    cur.execute("""
        SELECT u.telegram_id, u.personnel_id, u.verified_at, u.completed_year, u.completed_at,
               p.birthday, p.group_name, u.full_name
          FROM users u
          JOIN personnel p ON u.personnel_id = p.personnel_id
         WHERE u.telegram_id = ?
    """, (telegram_id,))
    r = cur.fetchone()
    if not r:
        return None
    return (
        r["telegram_id"], r["personnel_id"], r["verified_at"], r["completed_year"], r["completed_at"],
        r["birthday"], r["group_name"], r["full_name"]
    )

def get_deferment_by_pid(conn: sqlite3.Connection, personnel_id: str, year: int):
    cur = conn.cursor()
    cur.execute("SELECT reason, status FROM deferments WHERE personnel_id=? AND year=?", (personnel_id, year))
    return cur.fetchone()

# ---------- Commands ----------
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "Welcome to the IPPT Reminder Bot\n\n"
        "User:\n"
        "• /verify <PERSONNEL_ID> <YYYY-MM-DD>\n"
        "• /set_name <your name>\n"
        "• /status\n"
        "• /complete [YYYY-MM-DD] (within 100-day window)\n"
        "• /uncomplete (only during window)\n\n"
        "Admins: /admin_help"
    )

async def admin_help(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update.message.from_user.id):
        return await update.message.reply_text("Admins only.")
    await update.message.reply_text(
        "Admin commands:\n"
        "• /add_personnel <ID> <YYYY-MM-DD> [GROUP]\n"
        "• /update_birthday <PERSONNEL_ID> <YYYY-MM-DD>\n"
        "• /import_csv (then upload CSV/XLSX)\n"
        "• /report (All + per-group + Cycles_19_40; red=overdue, yellow=<100 days)\n"
        "• /defer_reason <tokens> [YEAR] -- <reason>\n"
        "• /defer_reset <tokens> [YEAR]\n"
        "• /admin_complete <tokens> [YEAR] [--date YYYY-MM-DD]\n"
        "• /admin_uncomplete <tokens> [YEAR]\n"
        "• /cycle_reason <tokens> [YEAR] -- <reason>\n"
        "• /cycle_reason_clear <tokens> [YEAR]\n"
        "• /unlink_user <tokens>\n"
        "• /remove_personnel <ID[,ID,...]>\n"
        "• /whoami"
    )

async def whoami(update: Update, context: ContextTypes.DEFAULT_TYPE):
    tid = update.message.from_user.id
    with closing(db_connect()) as conn:
        cur = conn.cursor()
        cur.execute("SELECT personnel_id, full_name FROM users WHERE telegram_id=?", (tid,))
        r = cur.fetchone()
    pid = r[0] if r and r[0] else "(not linked)"
    nm = r[1] if r and r[1] else "(no name set)"
    await update.message.reply_text(f"Telegram ID: {tid}\nLinked personnel_id: {pid}\nName: {nm}")

async def set_name(update: Update, context: ContextTypes.DEFAULT_TYPE):
    msg = update.message
    parts = msg.text.split(maxsplit=1)
    if len(parts) == 1:
        return await msg.reply_text("Usage: /set_name <your name>\nOr: /set_name --clear")
    payload = parts[1].strip()
    if payload in ("--clear", "-c"):
        with closing(db_connect()) as conn:
            cur = conn.cursor()
            cur.execute(
                "INSERT INTO users (telegram_id, full_name) VALUES (?, NULL) "
                "ON CONFLICT(telegram_id) DO UPDATE SET full_name=NULL",
                (msg.from_user.id,)
            )
            conn.commit()
        return await msg.reply_text("Name cleared.")
    name = payload[:120]
    with closing(db_connect()) as conn:
        cur = conn.cursor()
        cur.execute(
            "INSERT INTO users (telegram_id, full_name) VALUES (?, ?) "
            "ON CONFLICT(telegram_id) DO UPDATE SET full_name=excluded.full_name",
            (msg.from_user.id, name)
        )
        conn.commit()
    return await msg.reply_text(f"Name set to: {name}")

async def verify(update: Update, context: ContextTypes.DEFAULT_TYPE):
    parts = update.message.text.split()
    if len(parts) != 3:
        return await update.message.reply_text("Usage: /verify <PERSONNEL_ID> <YYYY-MM-DD>")
    pid = parts[1].strip()
    try:
        dob = parse_date_strict(parts[2].strip())
    except Exception:
        return await update.message.reply_text("Invalid date. Use YYYY-MM-DD.")
    with closing(db_connect()) as conn:
        cur = conn.cursor()
        cur.execute("SELECT birthday FROM personnel WHERE personnel_id=?", (pid,))
        r = cur.fetchone()
        if not r:
            return await update.message.reply_text("No such PERSONNEL_ID. Ask admin to add you.")
        if r[0] != format_date(dob):
            return await update.message.reply_text("Birthday does not match our records.")
        cur.execute(
            "INSERT INTO users (telegram_id, personnel_id, verified_at) VALUES (?, ?, ?) "
            "ON CONFLICT(telegram_id) DO UPDATE SET personnel_id=excluded.personnel_id, verified_at=excluded.verified_at",
            (update.message.from_user.id, pid, datetime.now(TZINFO).isoformat())
        )
        conn.commit()
    await update.message.reply_text("Verified and linked. Use /status.")

async def add_personnel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update.message.from_user.id):
        return await update.message.reply_text("Admins only.")
    parts = update.message.text.split()
    if len(parts) < 3:
        return await update.message.reply_text("Usage: /add_personnel <ID> <YYYY-MM-DD> [GROUP]")
    pid = parts[1].strip()
    try:
        dob = parse_date_strict(parts[2].strip())
    except Exception:
        return await update.message.reply_text("Invalid date. Use YYYY-MM-DD.")
    group = " ".join(parts[3:]).strip() if len(parts) > 3 else None
    with closing(db_connect()) as conn:
        cur = conn.cursor()
        cur.execute("INSERT OR REPLACE INTO personnel (personnel_id, birthday, group_name) VALUES (?, ?, ?)", (pid, format_date(dob), group))
        conn.commit()
    await update.message.reply_text(f"Added/updated {pid}.")

async def update_birthday(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update.message.from_user.id):
        return await update.message.reply_text("Admins only.")
    parts = update.message.text.split()
    if len(parts) != 3:
        return await update.message.reply_text("Usage: /update_birthday <PERSONNEL_ID> <YYYY-MM-DD>")
    pid = parts[1].strip()
    try:
        dob = parse_date_strict(parts[2].strip())
    except Exception:
        return await update.message.reply_text("Invalid date. Use YYYY-MM-DD.")
    with closing(db_connect()) as conn:
        cur = conn.cursor()
        cur.execute("UPDATE personnel SET birthday=? WHERE personnel_id=?", (format_date(dob), pid))
        if cur.rowcount == 0:
            return await update.message.reply_text("No such PERSONNEL_ID.")
        conn.commit()
    await update.message.reply_text(f"Updated {pid} birthday to {format_date(dob)}.")

# ---------- Import CSV/XLSX ----------
async def import_csv_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update.message.from_user.id):
        return await update.message.reply_text("Admins only.")
    context.user_data["awaiting_import"] = True
    await update.message.reply_text("Upload CSV/XLSX with headers: personnel_id,birthday[,group]. Extra columns ignored. (BOM-safe)")

async def document_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not context.user_data.get("awaiting_import"):
        return
    doc = update.message.document
    if not doc or not doc.file_name:
        return
    lower = doc.file_name.lower()
    if not (lower.endswith(".csv") or lower.endswith(".xlsx")):
        return await update.message.reply_text("Unsupported file type. Please upload .csv or .xlsx.")
    file = await doc.get_file()
    tmp_path = os.path.join("/tmp", doc.file_unique_id + ("_xlsx" if lower.endswith(".xlsx") else "_csv"))
    await file.download_to_drive(tmp_path)

    count = 0
    if lower.endswith(".csv"):
        with open(tmp_path, "rb") as f:
            data = f.read()
        if data.startswith(b"\xef\xbb\xbf"):
            data = data[3:]
        import io as _io
        reader = csv.DictReader(_io.StringIO(data.decode("utf-8", errors="replace")))
        for row in reader:
            pid = (row.get("personnel_id") or row.get("id") or "").strip()
            bday = (row.get("birthday") or row.get("dob") or "").strip()
            group = (row.get("group") or row.get("group_name") or "").strip() or None
            if not pid or not bday:
                continue
            try:
                dob = parse_date_strict(bday)
            except Exception:
                continue
            with closing(db_connect()) as conn:
                cur = conn.cursor()
                cur.execute("INSERT OR REPLACE INTO personnel (personnel_id, birthday, group_name) VALUES (?, ?, ?)", (pid, format_date(dob), group))
                conn.commit()
                count += 1
    else:
        from openpyxl import load_workbook
        wb = load_workbook(tmp_path)
        ws = wb.active
        headers = [str(c.value).strip().lower() if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
        def get_col(names):
            for i, h in enumerate(headers):
                if h in names:
                    return i
            return None
        pid_col = get_col({"personnel_id", "id"})
        bday_col = get_col({"birthday", "dob"})
        group_col = get_col({"group", "group_name"})
        if pid_col is None or bday_col is None:
            return await update.message.reply_text("XLSX must include 'personnel_id' and 'birthday' headers.")
        for row in ws.iter_rows(min_row=2):
            pid = str(row[pid_col].value).strip() if row[pid_col].value is not None else ""
            bday_cell = row[bday_col].value
            group = (str(row[group_col].value).strip() if (group_col is not None and row[group_col].value is not None) else None)
            if not pid or not bday_cell:
                continue
            try:
                if isinstance(bday_cell, datetime):
                    dob = bday_cell.date()
                else:
                    dob = parse_date_strict(str(bday_cell).strip())
            except Exception:
                continue
            with closing(db_connect()) as conn:
                cur = conn.cursor()
                cur.execute("INSERT OR REPLACE INTO personnel (personnel_id, birthday, group_name) VALUES (?, ?, ?)", (pid, format_date(dob), group))
                conn.commit()
                count += 1

    context.user_data["awaiting_import"] = False
    await update.message.reply_text(f"Imported {count} row(s).")

# ---------- Status ----------
async def status(update: Update, context: ContextTypes.DEFAULT_TYPE):
    msg = update.message
    today = current_local_date()
    tid = msg.from_user.id

    with closing(db_connect()) as conn:
        data = get_personnel_and_user(conn, tid)
    if not data:
        return await msg.reply_text("Please /verify first.")

    _, personnel_id, _, completed_year, _, birthday_str, group_name, full_name = data
    bday = parse_date_strict(birthday_str)

    in_window, start, end = today_in_window(bday, today)
    window_key = start.year
    next_start = adjusted_birthday_for_year(bday, start.year + 1)

    with closing(db_connect()) as conn:
        d = get_deferment_by_pid(conn, personnel_id, window_key)
    defer_reason, defer_status = (d[0], d[1]) if d else (None, None)

    cycle_start = start
    cycle_end_excl = adjusted_birthday_for_year(bday, start.year + 1)
    window_end = end

    with closing(db_connect()) as conn:
        cur = conn.cursor()
        cur.execute(
            "SELECT completed_at FROM completions WHERE telegram_id=? AND completed_at >= ? AND completed_at <= ? ORDER BY completed_at DESC LIMIT 1",
            (tid, iso_from_local_date(cycle_start, 0, 0), iso_from_local_date(window_end, 23, 59)),
        )
        row_win = cur.fetchone()
    completed_in_window_date = datetime.fromisoformat(row_win[0]).date() if row_win else None

    completed_in_cycle_date = None
    if not completed_in_window_date:
        with closing(db_connect()) as conn:
            cur = conn.cursor()
            cur.execute(
                "SELECT completed_at FROM completions WHERE telegram_id=? AND completed_at >= ? AND completed_at < ? ORDER BY completed_at DESC LIMIT 1",
                (tid, iso_from_local_date(cycle_start, 0, 0), iso_from_local_date(cycle_end_excl - timedelta(days=1), 23, 59)),
            )
            row_cyc = cur.fetchone()
        completed_in_cycle_date = datetime.fromisoformat(row_cyc[0]).date() if row_cyc else None

    if completed_in_window_date:
        window_result_line = f"100-day window result: ✅ Completed on time ({completed_in_window_date:%Y-%m-%d})"
    elif completed_in_cycle_date:
        overdue_days = (completed_in_cycle_date - window_end).days
        window_result_line = f"100-day window result: ⚠️ Completed overdue by {overdue_days} day(s) ({completed_in_cycle_date:%Y-%m-%d})"
    else:
        window_result_line = "100-day window result: ❌ Not completed"

    if defer_status == "approved":
        status_line = f"IPPT Status: Defer — {defer_reason}"
    elif in_window and completed_year == window_key:
        status_line = "IPPT Status: ✅ Completed"
    else:
        if today < start:
            status_line = f"IPPT Status: Window not open yet — starts {format_date(start)}"
        elif start <= today <= end:
            days_left = (end - today).days
            status_line = f"IPPT Status: {days_left} day(s) left to complete"
        else:
            status_line = f"IPPT Status: Window closed — next window starts {format_date(next_start)}"

    cycle_today_start, cycle_today_end_excl = cycle_for_date(bday, today)
    cycle_window_end_today = cycle_today_start + timedelta(days=WINDOW_DAYS)
    with closing(db_connect()) as conn:
        cur = conn.cursor()
        cur.execute(
            "SELECT completed_at FROM completions WHERE telegram_id=? AND completed_at >= ? AND completed_at < ? ORDER BY completed_at DESC LIMIT 1",
            (tid, iso_from_local_date(cycle_today_start, 0, 0), iso_from_local_date(cycle_today_end_excl - timedelta(days=1), 23, 59)),
        )
        r = cur.fetchone()
    if r:
        cd = datetime.fromisoformat(r[0]).date()
        if cd <= cycle_window_end_today:
            cycle_line = "Cycle status: ✅ Completed (on time)"
        else:
            cycle_line = f"Cycle status: ✅ Completed (overdue by {(cd - cycle_window_end_today).days} day(s))"
    else:
        cycle_line = "Cycle status: Not completed"

    lines = [
        status_line,
        f"Window: {format_date(start)} → {format_date(end)}",
        window_result_line,
        cycle_line,
        f"Today:  {format_date(today)}",
    ]
    if group_name:
        lines.append(f"Group:  {group_name}")
    if full_name:
        lines.append(f"Name:   {full_name}")
    lines.append(f"ID:     {personnel_id}")
    await msg.reply_text("\n".join(lines))

# ---------- User complete/uncomplete ----------
async def complete(update: Update, context: ContextTypes.DEFAULT_TYPE):
    msg = update.message
    today = current_local_date()
    parts = msg.text.split()
    given_date = None
    if len(parts) == 2:
        try:
            given_date = parse_date_strict(parts[1].strip())
        except Exception:
            return await msg.reply_text("Invalid date. Use YYYY-MM-DD, e.g., /complete 2025-01-10")

    with closing(db_connect()) as conn:
        data = get_personnel_and_user(conn, msg.from_user.id)
    if not data or not data[1]:
        return await msg.reply_text("You're not verified yet. Use /verify first.")

    _, _, _, _, _, birthday_str, _, _ = data
    bday = parse_date_strict(birthday_str)
    in_window, start, end = today_in_window(bday, today)
    if not in_window:
        return await msg.reply_text(f"You're outside your current window. Window: {format_date(start)} → {format_date(end)}")

    completion_date = today if given_date is None else given_date
    if not (start <= completion_date <= end):
        return await msg.reply_text(f"The supplied date must be within your current window {format_date(start)} to {format_date(end)}.")

    window_key = start.year
    now_iso = iso_from_local_date(completion_date, hour=9, minute=0)
    with closing(db_connect()) as conn:
        cur = conn.cursor()
        cur.execute("UPDATE users SET completed_year=?, completed_at=? WHERE telegram_id=?", (window_key, now_iso, msg.from_user.id))
        cur.execute("DELETE FROM completions WHERE telegram_id=? AND year=?", (msg.from_user.id, window_key))
        cur.execute("INSERT INTO completions (telegram_id, year, completed_at) VALUES (?, ?, ?)", (msg.from_user.id, window_key, now_iso))
        conn.commit()

    await msg.reply_text(
        f"Recorded as completed for the {WINDOW_DAYS}-day window starting {format_date(start)}.\n"
        f"(Window end: {format_date(end)}; date recorded: {format_date(completion_date)})"
    )

async def uncomplete(update: Update, context: ContextTypes.DEFAULT_TYPE):
    msg = update.message
    tid = msg.from_user.id
    today = current_local_date()

    with closing(db_connect()) as conn:
        data = get_personnel_and_user(conn, tid)
    if not data or not data[1]:
        return await msg.reply_text("You're not verified yet. Use /verify first.")

    _, _, _, _, _, birthday_str, _, _ = data
    bday = parse_date_strict(birthday_str)

    in_window, start, end = today_in_window(bday, today)
    if not in_window:
        return await msg.reply_text(
            "You can only use /uncomplete during your active 100-day window.\n"
            f"Current/next window: {format_date(start)} → {format_date(end)}\n"
            "If you need to clear a past window, please ask an admin to use /admin_uncomplete."
        )

    window_key = start.year

    with closing(db_connect()) as conn:
        cur = conn.cursor()
        cur.execute(
            "UPDATE users SET completed_year=NULL, completed_at=NULL WHERE telegram_id=? AND completed_year=?",
            (tid, window_key),
        )
        changed = cur.rowcount
        cur.execute("DELETE FROM completions WHERE telegram_id=? AND year=?", (tid, window_key))
        conn.commit()

    if changed:
        await msg.reply_text("Completion cleared for this window.")
    else:
        await msg.reply_text("There’s no recorded completion for this active window to clear.")

# ---------- Admin helpers ----------
async def _resolve_tokens_to_tids(tokens):
    tids = set()
    with closing(db_connect()) as conn:
        cur = conn.cursor()
        for t in tokens:
            t = t.strip()
            if not t:
                continue
            if t.isdigit():
                tids.add(int(t)); continue
            cur.execute("SELECT telegram_id FROM users WHERE personnel_id=?", (t,))
            for r in cur.fetchall():
                if r[0] is not None:
                    tids.add(int(r[0]))
    return tids

async def _resolve_tokens_to_pids(tokens):
    pids = set()
    with closing(db_connect()) as conn:
        cur = conn.cursor()
        for t in tokens:
            t = t.strip()
            if not t:
                continue
            cur.execute("SELECT 1 FROM personnel WHERE personnel_id=?", (t,))
            if cur.fetchone():
                pids.add(t); continue
            if t.isdigit():
                cur.execute("SELECT personnel_id FROM users WHERE telegram_id=?", (int(t),))
                r = cur.fetchone()
                if r and r[0]:
                    pids.add(r[0])
    return pids

# ---------- Admin: complete / uncomplete ----------
async def admin_complete(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update.message.from_user.id):
        return await update.message.reply_text("Admins only.")
    parts = update.message.text.split(maxsplit=1)
    if len(parts) < 2:
        return await update.message.reply_text(
            "Usage: /admin_complete <tokens> [YEAR] [--date YYYY-MM-DD]\n"
            "Also supports a bare YYYY-MM-DD without --date."
        )

    tail = parts[1].strip()
    m = re.search(r"--date\s+(\d{4}-\d{2}-\d{2})", tail)
    date_override = None
    if m:
        try:
            date_override = parse_date_strict(m.group(1))
        except Exception:
            return await update.message.reply_text("Invalid --date. Use YYYY-MM-DD.")
        tail = (tail[:m.start()] + tail[m.end():]).strip()

    tokens = [t for t in re.split(r"[,\s]+", tail) if t]

    if date_override is None:
        date_idx = next((i for i, t in enumerate(tokens) if re.fullmatch(r"\d{4}-\d{2}-\d{2}", t)), None)
        if date_idx is not None:
            try:
                date_override = parse_date_strict(tokens[date_idx])
            except Exception:
                return await update.message.reply_text("Invalid date. Use YYYY-MM-DD.")
            tokens.pop(date_idx)

    year = None
    if date_override is None and tokens and re.fullmatch(r"\d{4}", tokens[-1] or ""):
        year = int(tokens[-1])
        tokens = tokens[:-1]

    if not tokens:
        return await update.message.reply_text("No IDs provided.")

    tids = await _resolve_tokens_to_tids(tokens)
    if not tids:
        return await update.message.reply_text("No verified users matched these tokens.")

    updated = 0
    replaced = 0
    with closing(db_connect()) as conn:
        cur = conn.cursor()
        for tid in tids:
            cur.execute(
                "SELECT p.birthday FROM users u JOIN personnel p ON u.personnel_id=p.personnel_id WHERE u.telegram_id=?",
                (tid,),
            )
            r = cur.fetchone()
            if not r:
                continue
            bday = parse_date_strict(r[0])

            if date_override is not None:
                cstart, cend_excl = cycle_for_date(bday, date_override)
                if not (cstart <= date_override < cend_excl):
                    return await update.message.reply_text(
                        f"Date {format_date(date_override)} is not inside the birthday cycle for at least one user."
                    )
                target_year = cstart.year
                completion_iso = iso_from_local_date(date_override, hour=9, minute=0)
            else:
                if year is None:
                    _, start, _ = today_in_window(bday, current_local_date())
                    target_year = start.year
                else:
                    start = adjusted_birthday_for_year(bday, year)
                    target_year = start.year
                completion_iso = datetime.now(TZINFO).isoformat()

            cur.execute(
                "UPDATE users SET completed_year=?, completed_at=? WHERE telegram_id=?",
                (target_year, completion_iso, tid),
            )
            updated += cur.rowcount
            cur.execute("DELETE FROM completions WHERE telegram_id=? AND year=?", (tid, target_year))
            cur.execute(
                "INSERT INTO completions (telegram_id, year, completed_at) VALUES (?, ?, ?)",
                (tid, target_year, completion_iso),
            )
            replaced += 1
        conn.commit()

    note = " (date took precedence over YEAR)" if date_override is not None else ""
    return await update.message.reply_text(
        f"Admin completed. Users updated: {updated}, history rows replaced: {replaced}.{note}"
    )

async def admin_uncomplete(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update.message.from_user.id):
        return await update.message.reply_text("Admins only.")
    parts = update.message.text.split(maxsplit=1)
    if len(parts) < 2:
        return await update.message.reply_text("Usage: /admin_uncomplete <tokens> [YEAR]")
    tail = parts[1].strip()
    tokens = [t for t in re.split(r"[,\s]+", tail) if t]
    year = None
    if tokens and re.fullmatch(r"\d{4}", tokens[-1] or ""):
        year = int(tokens[-1]); tokens = tokens[:-1]
    if not tokens:
        return await update.message.reply_text("No IDs provided.")

    tids = await _resolve_tokens_to_tids(tokens)
    cleared = 0
    with closing(db_connect()) as conn:
        cur = conn.cursor()
        for tid in tids:
            if year is None:
                cur.execute("SELECT p.birthday FROM users u JOIN personnel p ON u.personnel_id=p.personnel_id WHERE u.telegram_id=?", (tid,))
                r = cur.fetchone()
                if not r:
                    continue
                bday = parse_date_strict(r[0])
                _, start, _ = today_in_window(bday, current_local_date())
                target_year = start.year
            else:
                target_year = year
            cur.execute("UPDATE users SET completed_year=NULL, completed_at=NULL WHERE telegram_id=? AND completed_year=?", (tid, target_year))
            cur.execute("DELETE FROM completions WHERE telegram_id=? AND year=?", (tid, target_year))
            cleared += cur.rowcount
        conn.commit()

    return await update.message.reply_text(f"Cleared completion for {cleared} user(s).")

# ---------- Admin: deferments & cycle reasons ----------
async def defer_reason(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update.message.from_user.id):
        return await update.message.reply_text("Admins only.")
    raw = update.message.text
    if " -- " not in raw:
        return await update.message.reply_text("Usage: /defer_reason <tokens> [YEAR] -- <reason>")
    head, reason = raw.split(" -- ", 1)
    parts = head.split(maxsplit=1)
    if len(parts) < 2:
        return await update.message.reply_text("Usage: /defer_reason <tokens> [YEAR] -- <reason>")
    tail = parts[1].strip()
    tokens = [t for t in re.split(r"[,\s]+", tail) if t]
    year = None
    if tokens and re.fullmatch(r"\d{4}", tokens[-1] or ""):
        year = int(tokens[-1]); tokens = tokens[:-1]
    if not tokens:
        return await update.message.reply_text("No IDs provided.")

    pids = await _resolve_tokens_to_pids(tokens)
    now = datetime.now(TZINFO).isoformat()
    with closing(db_connect()) as conn:
        cur = conn.cursor()
        for pid in pids:
            win_year = year
            if win_year is None:
                cur.execute("SELECT birthday FROM personnel WHERE personnel_id=?", (pid,))
                r = cur.fetchone()
                if not r:
                    continue
                bday = parse_date_strict(r[0])
                start, _ = window_for_date(bday, current_local_date())
                win_year = start.year
            cur.execute(
                "INSERT INTO deferments (personnel_id, year, reason, status, created_at) VALUES (?, ?, ?, 'approved', ?) "
                "ON CONFLICT(personnel_id, year) DO UPDATE SET reason=excluded.reason, status='approved'",
                (pid, win_year, reason.strip(), now),
            )
        conn.commit()
    await update.message.reply_text(f"Reason set for {len(pids)} user(s).")

async def defer_reset(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update.message.from_user.id):
        return await update.message.reply_text("Admins only.")
    parts = update.message.text.split(maxsplit=1)
    if len(parts) < 2:
        return await update.message.reply_text("Usage: /defer_reset <tokens> [YEAR]")
    tail = parts[1].strip()
    tokens = [t for t in re.split(r"[,\s]+", tail) if t]
    year = None
    if tokens and re.fullmatch(r"\d{4}", tokens[-1] or ""):
        year = int(tokens[-1]); tokens = tokens[:-1]
    if not tokens:
        return await update.message.reply_text("No IDs provided.")

    pids = await _resolve_tokens_to_pids(tokens)
    deleted = 0
    with closing(db_connect()) as conn:
        cur = conn.cursor()
        for pid in pids:
            win_year = year
            if win_year is None:
                cur.execute("SELECT birthday FROM personnel WHERE personnel_id=?", (pid,))
                r = cur.fetchone()
                if not r: continue
                bday = parse_date_strict(r[0])
                start, _ = window_for_date(bday, current_local_date())
                win_year = start.year
            cur.execute("DELETE FROM deferments WHERE personnel_id=? AND year=?", (pid, win_year))
            deleted += cur.rowcount
        conn.commit()
    return await update.message.reply_text(f"Deferments cleared: {deleted} row(s).")

async def cycle_reason(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update.message.from_user.id):
        return await update.message.reply_text("Admins only.")
    raw = update.message.text
    if " -- " not in raw:
        return await update.message.reply_text("Usage: /cycle_reason <tokens> [YEAR] -- <reason>")
    head, reason = raw.split(" -- ", 1)
    parts = head.split(maxsplit=1)
    if len(parts) < 2:
        return await update.message.reply_text("Usage: /cycle_reason <tokens> [YEAR] -- <reason>")
    tail = parts[1].strip()
    tokens = [t for t in re.split(r"[,\s]+", tail) if t]
    year = None
    if tokens and re.fullmatch(r"\d{4}", tokens[-1] or ""):
        year = int(tokens[-1]); tokens = tokens[:-1]
    if not tokens:
        return await update.message.reply_text("No IDs provided.")

    pids = await _resolve_tokens_to_pids(tokens)
    now = datetime.now(TZINFO).isoformat()
    with closing(db_connect()) as conn:
        cur = conn.cursor()
        for pid in pids:
            cyc_year = year
            if cyc_year is None:
                cur.execute("SELECT birthday FROM personnel WHERE personnel_id=?", (pid,))
                r = cur.fetchone()
                if not r: continue
                bday = parse_date_strict(r[0])
                start, _ = cycle_for_date(bday, current_local_date())
                cyc_year = start.year
            cur.execute(
                "INSERT INTO cycle_notes (personnel_id, year, reason, created_at) VALUES (?, ?, ?, ?) "
                "ON CONFLICT(personnel_id, year) DO UPDATE SET reason=excluded.reason",
                (pid, cyc_year, reason.strip(), now),
            )
        conn.commit()
    await update.message.reply_text(f"Cycle reason recorded for {len(pids)} user(s).")

async def cycle_reason_clear(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update.message.from_user.id):
        return await update.message.reply_text("Admins only.")
    parts = update.message.text.split(maxsplit=1)
    if len(parts) < 2:
        return await update.message.reply_text("Usage: /cycle_reason_clear <tokens> [YEAR]")
    tail = parts[1].strip()
    tokens = [t for t in re.split(r"[,\s]+", tail) if t]
    year = None
    if tokens and re.fullmatch(r"\d{4}", tokens[-1] or ""):
        year = int(tokens[-1]); tokens = tokens[:-1]
    if not tokens:
        return await update.message.reply_text("No IDs provided.")

    pids = await _resolve_tokens_to_pids(tokens)
    deleted = 0
    with closing(db_connect()) as conn:
        cur = conn.cursor()
        for pid in pids:
            cyc_year = year
            if cyc_year is None:
                cur.execute("SELECT birthday FROM personnel WHERE personnel_id=?", (pid,))
                r = cur.fetchone()
                if not r: continue
                bday = parse_date_strict(r[0])
                start, _ = cycle_for_date(bday, current_local_date())
                cyc_year = start.year
            cur.execute("DELETE FROM cycle_notes WHERE personnel_id=? AND year=?", (pid, cyc_year))
            deleted += cur.rowcount
        conn.commit()
    return await update.message.reply_text(f"Cycle reasons cleared: {deleted} row(s).")

# ---------- Admin: unlink & remove ----------
async def unlink_user(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update.message.from_user.id):
        return await update.message.reply_text("Admins only.")
    parts = update.message.text.split(maxsplit=1)
    if len(parts) < 2:
        return await update.message.reply_text("Usage: /unlink_user <tokens>")
    tokens = [t for t in re.split(r"[,\s]+", parts[1]) if t]

    cleared = 0
    with closing(db_connect()) as conn:
        cur = conn.cursor()
        for t in tokens:
            if t.isdigit():
                tid = int(t)
                cur.execute("UPDATE users SET personnel_id=NULL, verified_at=NULL, completed_year=NULL, completed_at=NULL WHERE telegram_id=?", (tid,))
                cleared += cur.rowcount
        for t in tokens:
            if not t.isdigit():
                cur.execute("UPDATE users SET personnel_id=NULL, verified_at=NULL, completed_year=NULL, completed_at=NULL WHERE personnel_id=?", (t,))
                cleared += cur.rowcount
        conn.commit()

    await update.message.reply_text(f"Unlinked {cleared} mapping(s).")

async def remove_personnel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update.message.from_user.id):
        return await update.message.reply_text("Admins only.")
    parts = update.message.text.split(maxsplit=1)
    if len(parts) < 2:
        return await update.message.reply_text("Usage: /remove_personnel <ID[,ID,...]>")
    tokens = [t.strip() for t in re.split(r"[,\s]+", parts[1]) if t.strip()]

    removed = 0
    with closing(db_connect()) as conn:
        cur = conn.cursor()
        for pid in tokens:
            cur.execute("UPDATE users SET personnel_id=NULL, verified_at=NULL, completed_year=NULL, completed_at=NULL WHERE personnel_id=?", (pid,))
            cur.execute("DELETE FROM deferments WHERE personnel_id=?", (pid,))
            cur.execute("DELETE FROM cycle_notes WHERE personnel_id=?", (pid,))
            cur.execute("DELETE FROM personnel WHERE personnel_id=?", (pid,))
            removed += cur.rowcount
        conn.commit()

    await update.message.reply_text(f"Removed {removed} personnel record(s).")

# ---------- Reports ----------
async def report(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update.message.from_user.id):
        return await update.message.reply_text("Admins only.")
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    today = current_local_date()

    with closing(db_connect()) as conn:
        cur = conn.cursor()
        cur.execute("""
            SELECT p.personnel_id, p.birthday, p.group_name,
                   u.telegram_id, u.completed_year, u.completed_at, u.full_name
              FROM personnel p
              LEFT JOIN users u ON p.personnel_id = u.personnel_id
        """)
        rows = cur.fetchall()

    RED_FILL    = PatternFill(start_color="FFFFC0C0", end_color="FFFFC0C0", fill_type="solid")
    YELLOW_FILL = PatternFill(start_color="FFFFFF99", end_color="FFFFFF99", fill_type="solid")

    def build_current_row(pid, bday_str, group_name, telegram_id, completed_year, completed_at, full_name):
        bday = parse_date_strict(bday_str)
        _, start, end = today_in_window(bday, today)
        window_key = start.year
        done = (completed_year == window_key)
        verified = bool(telegram_id)

        with closing(db_connect()) as conn2:
            d = get_deferment_by_pid(conn2, pid, window_key)
        d_status = (d[1] if d else "")
        d_reason = (d[0] if d else "")

        cyc_start, cyc_end_excl = cycle_for_date(bday, today)
        cyc_win_end = cyc_start + timedelta(days=WINDOW_DAYS)

        completed_at_str = None
        if telegram_id:
            with closing(db_connect()) as conn2:
                c2 = conn2.cursor()
                c2.execute(
                    "SELECT completed_at FROM completions WHERE telegram_id=? AND completed_at >= ? AND completed_at < ? ORDER BY completed_at DESC LIMIT 1",
                    (telegram_id, iso_from_local_date(cyc_start, 0, 0), iso_from_local_date(cyc_end_excl - timedelta(days=1), 23, 59)),
                )
                r = c2.fetchone()
                if r:
                    completed_at_str = r[0]

        cycle_status = "not_completed"
        cycle_overdue_days = ""
        if completed_at_str:
            cd = datetime.fromisoformat(completed_at_str).date()
            if cd <= cyc_win_end:
                cycle_status = "on_time"
            else:
                cycle_status = "overdue"
                cycle_overdue_days = (cd - cyc_win_end).days
        else:
            with closing(db_connect()) as conn2:
                c2 = conn2.cursor()
                c2.execute("SELECT reason FROM cycle_notes WHERE personnel_id=? AND year=?", (pid, cyc_start.year))
                note = c2.fetchone()
            if note and note[0]:
                cycle_status = f"not_completed ({note[0]})"

        if done:
            days_left, days_overdue = "", ""
        else:
            if today <= end:
                days_left, days_overdue = (end - today).days, ""
            else:
                days_left, days_overdue = "", (today - end).days

        highlight_red = (d_status != "approved") and (not done) and (today > end)
        highlight_yellow = (d_status != "approved") and (not done) and (start <= today <= end) and ((end - today).days < 100)

        return {
            "personnel_id": pid,
            "name": full_name or "",
            "birthday": bday_str,
            "group_name": group_name or "",
            "verified": "yes" if verified else "no",
            "window_start": format_date(start),
            "window_end": format_date(end),
            "completed_this_window": "yes" if done else "no",
            "completed_at": completed_at or "",
            "deferment_status": d_status or "",
            "deferment_reason": d_reason or "",
            "days_left": days_left,
            "days_overdue": days_overdue,
            "cycle_status": cycle_status,
            "cycle_overdue_days": cycle_overdue_days,
            "_highlight_red": highlight_red,
            "_highlight_yellow": highlight_yellow,
        }

    data_rows = [build_current_row(*r) for r in rows]

    headers_all = [
        "personnel_id","name","birthday","group_name","verified",
        "window_start","window_end",
        "completed_this_window","completed_at",
        "deferment_status","deferment_reason",
        "days_left","days_overdue",
        "cycle_status","cycle_overdue_days"
    ]

    wb = Workbook()
    ws_all = wb.active
    ws_all.title = "All"
    ws_all.append(headers_all)
    for c in ws_all[1]:
        c.font = Font(bold=True)

    def write_row(ws, rec):
        ws.append([rec[h] for h in headers_all])
        fill = None
        if rec.get("_highlight_red"):
            fill = RED_FILL
        elif rec.get("_highlight_yellow"):
            fill = YELLOW_FILL
        if fill:
            for cell in ws[ws.max_row]:
                cell.fill = fill

    for rec in data_rows:
        write_row(ws_all, rec)

    for col_idx in range(1, ws_all.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 12
        for row in range(1, ws_all.max_row + 1):
            val = ws_all.cell(row=row, column=col_idx).value
            max_len = max(max_len, len(str(val)) if val is not None else 0)
        ws_all.column_dimensions[col_letter].width = min(40, max_len + 2)

    # Per-group sheets
    groups = {}
    for rec in data_rows:
        key = (rec["group_name"] or "No Group")
        groups.setdefault(key, []).append(rec)

    def safe_sheet_name(name: str) -> str:
        bad = ["\", "/", "?", "*", "[", "]"]
        for b in bad:
            name = name.replace(b, " ")
        name = name.strip() or "No Group"
        return name[:31]

    for gname, recs in sorted(groups.items(), key=lambda kv: kv[0].lower()):
        ws = wb.create_sheet(title=safe_sheet_name(gname))
        ws.append(headers_all)
        for c in ws[1]:
            c.font = Font(bold=True)
        for rec in recs:
            write_row(ws, rec)
        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            max_len = 12
            for row in range(1, ws.max_row + 1):
                val = ws.cell(row=row, column=col_idx).value
                max_len = max(max_len, len(str(val)) if val is not None else 0)
            ws.column_dimensions[col_letter].width = min(40, max_len + 2)

    # Cycles 19–40
    ws_cyc = wb.create_sheet(title="Cycles_19_40")
    headers_cyc = [
        "personnel_id","name","group_name","age","cycle_start","cycle_end",
        "window_end","verified","status","overdue_days","completed_at","note"
    ]
    ws_cyc.append(headers_cyc)
    for c in ws_cyc[1]:
        c.font = Font(bold=True)

    with closing(db_connect()) as conn:
        cur = conn.cursor()
        cur.execute("SELECT telegram_id, completed_at FROM completions")
        compl = {}
        for tid, iso in cur.fetchall():
            if tid is None or not iso:
                continue
            compl.setdefault(int(tid), []).append((datetime.fromisoformat(iso).date(), iso))

    with closing(db_connect()) as conn:
        cur = conn.cursor()
        cur.execute("SELECT p.personnel_id, p.birthday, p.group_name, u.telegram_id, u.full_name FROM personnel p LEFT JOIN users u ON p.personnel_id=u.personnel_id")
        pers = cur.fetchall()

    for pid, bday_str, group_name, telegram_id, full_name in pers:
        bday = parse_date_strict(bday_str)
        birth_year = bday.year
        for age in range(19, 41):
            cyc_start = adjusted_birthday_for_year(bday, birth_year + age)
            cyc_end_excl = adjusted_birthday_for_year(bday, birth_year + age + 1)
            window_end = cyc_start + timedelta(days=WINDOW_DAYS)

            status = "not_completed"
            overdue_days = ""
            completed_at_out = ""
            note = ""

            if telegram_id and int(telegram_id) in compl:
                best = None
                for d, iso in compl[int(telegram_id)]:
                    if cyc_start <= d < cyc_end_excl:
                        if best is None or d > best[0]:
                            best = (d, iso)
                if best:
                    completed_at_out = best[0].strftime("%Y-%m-%d")
                    if best[0] <= window_end:
                        status = "on_time"
                    else:
                        status = "overdue"
                        overdue_days = (best[0] - window_end).days
            if status == "not_completed":
                with closing(db_connect()) as conn2:
                    c2 = conn2.cursor()
                    c2.execute("SELECT reason FROM cycle_notes WHERE personnel_id=? AND year=?", (pid, cyc_start.year))
                    r = c2.fetchone()
                    if r and r[0]:
                        note = r[0]

            ws_cyc.append([
                pid, (full_name or ""), (group_name or ""), age,
                format_date(cyc_start), format_date(cyc_end_excl - timedelta(days=1)),
                format_date(window_end),
                "yes" if telegram_id else "no",
                status, overdue_days, completed_at_out, note
            ])

    for col_idx in range(1, ws_cyc.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 12
        for row in range(1, ws_cyc.max_row + 1):
            val = ws_cyc.cell(row=row, column=col_idx).value
            max_len = max(max_len, len(str(val)) if val is not None else 0)
        ws_cyc.column_dimensions[col_letter].width = min(50, max_len + 2)

    out = io.BytesIO()
    wb.save(out); out.seek(0)
    await update.message.reply_document(document=InputFile(out, filename="ippt_100day_report.xlsx"),
                                        caption="Report: All + per-group + Cycles_19_40 (Name included)")


# ---------- Audit ----------
async def defer_audit(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update.message.from_user.id):
        return await update.message.reply_text("Admins only.")
    out = io.StringIO()
    writer = csv.writer(out)
    writer.writerow(["personnel_id","year","reason","status","created_at"])
    with closing(db_connect()) as conn:
        cur = conn.cursor()
        cur.execute("SELECT personnel_id, year, reason, status, created_at FROM deferments ORDER BY year, personnel_id")
        for pid, yr, reason, status, created_at in cur.fetchall():
            writer.writerow([pid, yr, reason or "", status or "", created_at or ""])
    data = out.getvalue().encode("utf-8")
    bio = io.BytesIO(data); bio.seek(0)
    await update.message.reply_document(document=InputFile(bio, filename="deferment_audit.csv"),
                                        caption="Deferment audit CSV")

# ---------- Scheduler ----------
async def daily_reminder_job(context: ContextTypes.DEFAULT_TYPE):
    today = current_local_date()
    with closing(db_connect()) as conn:
        cur = conn.cursor()
        cur.execute("""
            SELECT u.telegram_id, u.personnel_id, p.birthday, u.completed_year
              FROM users u
              JOIN personnel p ON u.personnel_id = p.personnel_id
        """)
        rows = cur.fetchall()

    for telegram_id, pid, bday_str, completed_year in rows:
        bday = parse_date_strict(bday_str)
        in_window, start, end = today_in_window(bday, today)
        window_key = start.year

        with closing(db_connect()) as conn:
            d = get_deferment_by_pid(conn, pid, window_key)
        skip = bool(d and d[1] == "approved")
        done = (completed_year == window_key)

        # end-of-window maintenance
        if today > end:
            try:
                with closing(db_connect()) as conn2:
                    c2 = conn2.cursor()
                    c2.execute("DELETE FROM deferments WHERE personnel_id=? AND year=?", (pid, window_key))
                    conn2.commit()
            except Exception:
                pass
            try:
                with closing(db_connect()) as conn3:
                    c3 = conn3.cursor()
                    c3.execute("UPDATE users SET completed_year=NULL, completed_at=NULL WHERE telegram_id=? AND completed_year=?", (telegram_id, window_key))
                    conn3.commit()
            except Exception:
                pass

        if in_window and not done and not skip:
            days_since_start = (today - start).days
            if days_since_start % REMINDER_INTERVAL_DAYS == 0:
                try:
                    await context.bot.send_message(
                        chat_id=telegram_id,
                        text=(
                            f"Reminder: Your IPPT window is {format_date(start)} → {format_date(end)}.\n"
                            f"Use /complete when done. You can also do /status anytime."
                        ),
                    )
                except Exception:
                    pass

# ---------- Wiring ----------
def setup_handlers(app):
    # User
    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("whoami", whoami))
    app.add_handler(CommandHandler("set_name", set_name))
    app.add_handler(CommandHandler("verify", verify))
    app.add_handler(CommandHandler("status", status))
    app.add_handler(CommandHandler("complete", complete))
    app.add_handler(CommandHandler("uncomplete", uncomplete))

    # Admin
    app.add_handler(CommandHandler("admin_help", admin_help))
    app.add_handler(CommandHandler("add_personnel", add_personnel))
    app.add_handler(CommandHandler("update_birthday", update_birthday))
    app.add_handler(CommandHandler("import_csv", import_csv_cmd))
    app.add_handler(CommandHandler("report", report))
    app.add_handler(CommandHandler("defer_reason", defer_reason))
    app.add_handler(CommandHandler("defer_reset", defer_reset))
    app.add_handler(CommandHandler("admin_complete", admin_complete))
    app.add_handler(CommandHandler("admin_uncomplete", admin_uncomplete))
    app.add_handler(CommandHandler("cycle_reason", cycle_reason))
    app.add_handler(CommandHandler("cycle_reason_clear", cycle_reason_clear))
    app.add_handler(CommandHandler("unlink_user", unlink_user))
    app.add_handler(CommandHandler("remove_personnel", remove_personnel))
    app.add_handler(CommandHandler("defer_audit", defer_audit))

    # File uploads for /import_csv
    app.add_handler(MessageHandler(filters.Document.ALL & (~filters.COMMAND), document_handler))

def schedule_jobs(app):
    app.job_queue.run_daily(daily_reminder_job, time=time(hour=9, minute=0, tzinfo=TZINFO), name="daily_reminders")

def main():
    init_db()
    if not BOT_TOKEN:
        raise SystemExit("Missing BOT_TOKEN env var.")
    app = ApplicationBuilder().token(BOT_TOKEN).build()
    setup_handlers(app)
    schedule_jobs(app)
    print("Bot is running…")
    app.run_polling(close_loop=False)

if __name__ == "__main__":
    main()
